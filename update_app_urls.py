"""
Run this script once to replace app_url_reference table
from the Sites Excel file.

Usage:
    python update_app_urls.py
"""

import io, sys
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

XLSX = r"C:\Users\HP\Desktop\$Screenshot\Fianl_Site_for_automation_standardized.xlsx"
DB   = "postgresql://postgres:Arun%40123%24@localhost:5432/ctr_db"

print("Connecting to database...")
conn = psycopg2.connect(DB)
cur  = conn.cursor()

print(f"Reading Excel: {XLSX}\n")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
results = []

for sheet_name in wb.sheetnames:
    if sheet_name.strip().lower() == "summary":
        print(f"  SKIP (summary): {sheet_name}")
        continue

    ws        = wb[sheet_name]
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        continue

    data_rows = rows_iter[1:]   # skip header row
    boxes     = [(0, 1), (4, 5), (8, 9)]  # (id_col, url_col) pairs
    records   = []

    for row in data_rows:
        for id_ci, url_ci in boxes:
            uid     = row[id_ci]  if len(row) > id_ci  else None
            url     = row[url_ci] if len(row) > url_ci else None
            if url is None:
                continue
            url_str = str(url).strip()
            if not url_str or url_str.lower() in ("none", "nan", "sites"):
                continue
            records.append((sheet_name.strip(), int(uid) if uid else None, url_str))

    cur.execute("DELETE FROM app_url_reference WHERE sheet_name = %s", (sheet_name.strip(),))
    if records:
        execute_values(
            cur,
            "INSERT INTO app_url_reference (sheet_name, url_id, url) VALUES %s",
            records,
            page_size=500,
        )
    conn.commit()
    results.append((sheet_name.strip(), len(records)))
    print(f"  ✓ {sheet_name.strip():35s} → {len(records)} URLs")

wb.close()
cur.close()
conn.close()

total = sum(r[1] for r in results)
print(f"\n✅ Done — {len(results)} sheets updated, {total} total URLs inserted.")
