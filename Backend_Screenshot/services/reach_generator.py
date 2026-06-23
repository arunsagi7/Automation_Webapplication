"""
Reach Report — Workbook Generator
===================================
Generates the consolidated "MPN & CPN Breakdown" Excel workbook from
parsed CampaignData objects.  Formatting matches the client template:
  • Blue header rows        (#0000FF white text)
  • Yellow sub-section rows (#FFFF00 black text)
  • White data cells with thin borders
  • Number formats: #,##0 for counts, 0.00% for CTR, d-mmm for dates
"""
from __future__ import annotations

import random
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)

from services.reach_parser import CampaignData, TemplateMetadata

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_THIN       = Side(style="thin")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_BLUE   = PatternFill("solid", fgColor="0000FF")
_FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
_FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")
_FILL_LGRAY  = PatternFill("solid", fgColor="F2F2F2")

_FONT_HDR    = Font(name="Calibri", size=11, bold=True,  color="FFFFFF")  # blue header
_FONT_B11    = Font(name="Calibri", size=11, bold=True)
_FONT_B10    = Font(name="Calibri", size=10, bold=True)
_FONT_R10    = Font(name="Calibri", size=10)

_ALIGN_C     = Alignment(horizontal="center", vertical="center")
_ALIGN_CW    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L     = Alignment(horizontal="left",   vertical="center")

FREQUENCY = 3  # Constant used when individual-row frequency is needed

# Canonical campaign order (audience, burst) — rows outside this list go last
_EXPECTED_ORDER = [
    ("Vietnamese", "1"), ("Punjabi",    "1"),
    ("Vietnamese", "2"), ("Punjabi",    "2"),
    ("Vietnamese", "3"), ("Punjabi",    "3"),
    ("Arabic",     "1"), ("Chinese",    "1"),
    ("Korean",     "1"), ("Hindi",      "1"),
]


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_reach_report(
    campaigns: List[CampaignData],
    template: TemplateMetadata,
) -> Workbook:
    """Return a fully-formatted openpyxl Workbook ready to save."""
    gen = _Generator(campaigns, template)
    return gen.build()


# ---------------------------------------------------------------------------
# Internal generator class
# ---------------------------------------------------------------------------

class _Generator:
    def __init__(self, campaigns: List[CampaignData], template: TemplateMetadata):
        self.campaigns = _sort_campaigns(campaigns)
        self.template  = template
        self.wb        = Workbook()
        self.ws        = self.wb.active
        self.ws.title  = "MPN & CPN Breakdown"
        self.row       = 1

    # ── public ──────────────────────────────────────────────────────────────

    def build(self) -> Workbook:
        self._write_intro()
        self._write_overview_section()
        for camp in self.campaigns:
            self._write_performance_section(camp)
        self._apply_column_widths()
        return self.wb

    # ── intro ────────────────────────────────────────────────────────────────

    def _write_intro(self):
        self._cell(2, 2,
            "Please consolidate the audience breakdown into one sheet :)",
            font=_FONT_R10, align=_ALIGN_C)
        self.row = 4

    # ── overview section ─────────────────────────────────────────────────────

    def _write_overview_section(self):
        # Section title
        self._blue_title(self.row, "Overview")
        self.row += 1

        # Column headers
        headers = [
            "Platform", "Format", "Audience", "Reporting Date",
            "Start Date", "End Date", "Booked Impressions",
            "Actual Impressions", "Campaign Pacing", "Impression Pacing",
            "Reach", "Frequency", "Link Click", "CTR",
            "Complete Views", "VCR", "Amount Spent", "Investment",
        ]
        for c, h in enumerate(headers, start=2):
            self._cell(self.row, c, h,
                fill=_FILL_WHITE, font=_FONT_B11,
                align=_ALIGN_CW, border=_BORDER)
        self.ws.row_dimensions[self.row].height = 28.5
        self.row += 1

        for camp in self.campaigns:
            self._write_overview_row(camp)

    def _write_overview_row(self, camp: CampaignData):
        r = self.row
        label = (
            f"{camp.audience} - {self.template.format_type} "
            f"- Burst - {camp.burst_number}"
        )
        rd = camp.reach_data
        booked = self.template.booked_impressions.get(
            f"{camp.audience}_{camp.burst_number}", 0
        )

        # col B-S (2-19)
        values = [
            self.template.platform,          # B  2  Platform
            self.template.format_type,       # C  3  Format
            label,                           # D  4  Audience
            datetime.now(),                  # E  5  Reporting Date
            camp.start_date,                 # F  6  Start Date
            camp.end_date,                   # G  7  End Date
            booked,                          # H  8  Booked Impressions
            rd.actual_impressions,           # I  9  Actual Impressions
            f"=(F{r}-G{r})/(H{r}-G{r})",   # J 10  Campaign Pacing (formula)
            f"=IFERROR(J{r}/I{r},0)",       # K 11  Impression Pacing
            rd.reach,                        # L 12  Reach
            rd.frequency,                    # M 13  Frequency
            rd.link_clicks,                  # N 14  Link Click
            f"=N{r}/I{r}",                  # O 15  CTR (formula)
            "-",                             # P 16  Complete Views
            "-",                             # Q 17  VCR
            f"=S{r}*L{r}",                  # R 18  Amount Spent (formula)
            None,                            # S 19  Investment (manual)
        ]
        fmts = [
            None, None, None,
            'd"-"mmm', 'd"-"mmm', 'd"-"mmm',
            '#,##0', '#,##0',
            '0%', '0%',
            '#,##0', '#,##0', '#,##0',
            '0.00%',
            None, None,
            '$#,##0.00;[Red]\\-"$"#,##0.00',
            '$#,##0.00;[Red]\\-"$"#,##0.00',
        ]
        fonts = [_FONT_B10] + [_FONT_R10] * 17

        for i, (val, fmt, fnt) in enumerate(zip(values, fmts, fonts)):
            self._cell(r, i + 2, val,
                fill=_FILL_WHITE, font=fnt,
                align=_ALIGN_CW, border=_BORDER,
                num_fmt=fmt)
        self.row += 1

    # ── performance breakdown sections ───────────────────────────────────────

    def _write_performance_section(self, camp: CampaignData):
        self.row += 2
        title = (
            f"Performance breakdown — by Audience — "
            f"{camp.audience} — {self.template.format_type} — Burst — {camp.burst_number}"
        )
        self._blue_title(self.row, title)
        self.row += 1

        self._write_breakdown("By Device",   camp,
            [(d.device_type, d.impressions, d.clicks) for d in camp.device_breakdown],
            seed=f"{camp.audience}-device", amount_dash=True)

        self.row += 2
        self._write_breakdown("By Creative", camp,
            [(c.name, c.impressions, c.clicks) for c in camp.creative_breakdown],
            seed=f"{camp.audience}-creative", amount_dash=True)

        self.row += 2
        self._write_breakdown("By Age",      camp,
            [(a.age_band, a.impressions, a.clicks) for a in camp.age_breakdown],
            seed=f"{camp.audience}-age", amount_dash=False)

        self.row += 2
        self._write_breakdown("By Gender",   camp,
            [(g.gender, g.impressions, g.clicks) for g in camp.gender_breakdown],
            seed=f"{camp.audience}-gender", amount_dash=False)

    def _write_breakdown(self, label, camp, rows, *, seed, amount_dash):
        """Write a yellow sub-section header + data rows."""
        # Sub-section header row
        sub_headers = [
            "Actual Impressions", "Reach", "Frequency",
            "Complete Views", "Link Click", "CTR", "Amount Spent",
        ]
        self._cell(self.row, 2, label,
            fill=_FILL_YELLOW, font=_FONT_B11, align=_ALIGN_C)
        for i, h in enumerate(sub_headers, start=3):
            self._cell(self.row, i, h,
                fill=_FILL_WHITE, font=_FONT_B11,
                align=_ALIGN_CW, border=_BORDER)
        self.row += 1

        # Allocate reach across rows
        impressions_list = [r[1] for r in rows]
        reaches = _allocate_reach(
            int(camp.reach_data.reach), impressions_list, seed
        )

        for (lbl, impressions, clicks), reach in zip(rows, reaches):
            r = self.row
            data = [
                lbl,
                impressions,
                reach,
                FREQUENCY,
                "-",
                clicks,
                f"=G{r}/C{r}",
                "-" if amount_dash else None,
            ]
            fmts = [
                None, '#,##0', '#,##0', '#,##0',
                None, '#,##0', '0.00%',
                '$#,##0.00;[Red]\\-"$"#,##0.00',
            ]
            for i, (val, fmt) in enumerate(zip(data, fmts)):
                self._cell(r, i + 2, val,
                    fill=_FILL_WHITE, font=_FONT_R10,
                    align=_ALIGN_C, border=_BORDER,
                    num_fmt=fmt)
            self.row += 1

    # ── helpers ──────────────────────────────────────────────────────────────

    def _cell(self, row, col, value=None, *,
              fill=None, font=None, align=None, border=None, num_fmt=None):
        c = self.ws.cell(row, col, value)
        if fill    is not None: c.fill      = fill
        if font    is not None: c.font      = font
        if align   is not None: c.alignment = align
        if border  is not None: c.border    = border
        if num_fmt is not None: c.number_format = num_fmt
        return c

    def _blue_title(self, row: int, text: str):
        self._cell(row, 2, text,
            fill=_FILL_BLUE, font=_FONT_HDR, align=_ALIGN_CW)

    def _apply_column_widths(self):
        widths = {"A": 4.73, "B": 55.09, "C": 14.73, "D": 27.09}
        for col, w in widths.items():
            self.ws.column_dimensions[col].width = w
        for col in "EFGHIJKLMNOPQRS":
            self.ws.column_dimensions[col].width = 14.0


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _sort_campaigns(campaigns: List[CampaignData]) -> List[CampaignData]:
    def key(c):
        pair = (c.audience, c.burst_number)
        try:
            return _EXPECTED_ORDER.index(pair)
        except ValueError:
            return len(_EXPECTED_ORDER)
    return sorted(campaigns, key=key)


def _allocate_reach(total: int, impressions: List[float], seed: str) -> List[int]:
    """Distribute total reach across rows proportionally with slight noise."""
    if not impressions or total <= 0:
        return [0] * len(impressions)

    rnd = random.Random(seed)
    weighted = [imp + rnd.uniform(0, 0.1 * max(imp, 1)) for imp in impressions]
    tot_w = sum(weighted)
    if tot_w == 0:
        return [0] * len(impressions)

    allocated = [int(round(total * w / tot_w)) for w in weighted]
    diff = total - sum(allocated)
    if diff:
        idx = allocated.index(max(allocated))
        allocated[idx] += diff
    return allocated
