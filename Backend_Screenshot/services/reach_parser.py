"""
Reach Report — Excel Parser & Data Models
==========================================
Parses individual campaign burst files (Vietnamese/Punjabi - Burst N) and
a master template file into structured dataclasses for the report generator.

Supported audiences:  Vietnamese, Punjabi, Arabic, Chinese, Korean, Hindi,
                      Tamil, Cantonese, Greek, Italian, Spanish (extend as needed)
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class ReachData:
    actual_impressions: float
    link_clicks: float
    ctr: float
    reach: float
    frequency: float


@dataclass
class DeviceBreakdown:
    device_type: str
    impressions: float
    clicks: float
    ctr: float


@dataclass
class CreativeBreakdown:
    name: str
    impressions: float
    clicks: float
    ctr: float


@dataclass
class AgeBreakdown:
    age_band: str
    impressions: float
    clicks: float
    ctr: float


@dataclass
class GenderBreakdown:
    gender: str
    impressions: float
    clicks: float
    ctr: float


@dataclass
class CampaignData:
    audience: str
    burst_number: str
    reach_data: ReachData
    device_breakdown: List[DeviceBreakdown]
    creative_breakdown: List[CreativeBreakdown]
    age_breakdown: List[AgeBreakdown]
    gender_breakdown: List[GenderBreakdown]
    start_date: Optional[datetime]
    end_date: Optional[datetime]


@dataclass
class TemplateMetadata:
    platform: str
    format_type: str
    booked_impressions: Dict[str, float] = field(default_factory=dict)
    start_date: Dict[str, str] = field(default_factory=dict)
    end_date: Dict[str, str] = field(default_factory=dict)
    reporting_date: Dict[str, str] = field(default_factory=dict)
    audience_labels: Dict[str, str] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Known audiences — extend this list as needed
# ---------------------------------------------------------------------------
_AUDIENCE_KEYWORDS = [
    "vietnamese", "punjabi", "arabic", "chinese", "korean",
    "hindi", "tamil", "cantonese", "greek", "italian",
    "spanish", "mandarin", "urdu", "bengali", "turkish",
]


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

REQUIRED_SHEETS = [
    "REACH", "DATE", "APP URL", "TIME OF DAY",
    "EXCHANGE", "DEVICE", "CREATIVE", "CITY", "AGE", "GENDER",
]

DEVICE_MAPPING = {
    "Mobile": "Mobile",
    "Smart Phone": "Mobile",
    "Smartphone": "Mobile",
    "Tablet": "Tablet",
    "Desktop": "Desktop",
    "Connected TV": "Connected TV",
    "CTV": "Connected TV",
}


def parse_filename(filename: str) -> Tuple[str, str]:
    """
    Extract audience and burst number from a campaign filename.
    Returns (audience_title_case, burst_number_str).

    Accepts patterns like:
      Vietnamese_Burst_1.xlsx
      Punjabi-burst-2_report.xlsx
      Arabic Burst 3.xlsx
    """
    name_lower = filename.lower()

    audience = None
    for kw in _AUDIENCE_KEYWORDS:
        if kw in name_lower:
            audience = kw.title()
            break

    if audience is None:
        # Fallback: use whatever is before "burst" in the filename
        m = re.search(r"([a-zA-Z]+)[_\-\s]+burst", name_lower)
        if m:
            audience = m.group(1).title()
        else:
            raise ValueError(
                f"Cannot determine audience from filename: {filename!r}. "
                f"Expected one of: {', '.join(k.title() for k in _AUDIENCE_KEYWORDS)}"
            )

    burst_match = re.search(r"burst[_\-\s]*(\d+)", name_lower)
    if not burst_match:
        raise ValueError(
            f"Cannot find burst number in filename: {filename!r}. "
            "Expected 'BurstN' or 'Burst_N' or 'Burst N'."
        )
    burst_number = burst_match.group(1)

    return audience, burst_number


def parse_campaign_file(filepath: str) -> CampaignData:
    """Parse a single campaign burst Excel file into CampaignData."""
    wb = load_workbook(filepath, data_only=True)

    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"Missing required sheet(s): {', '.join(missing)}")

    filename = filepath.replace("\\", "/").split("/")[-1]
    audience, burst_number = parse_filename(filename)

    reach_data        = _parse_reach_sheet(wb["REACH"])
    device_breakdown  = _parse_device_sheet(wb["DEVICE"])
    creative_breakdown= _parse_creative_sheet(wb["CREATIVE"])
    age_breakdown     = _parse_age_sheet(wb["AGE"])
    gender_breakdown  = _parse_gender_sheet(wb["GENDER"])
    start_date, end_date = _parse_date_sheet(wb["DATE"])

    return CampaignData(
        audience=audience,
        burst_number=burst_number,
        reach_data=reach_data,
        device_breakdown=device_breakdown,
        creative_breakdown=creative_breakdown,
        age_breakdown=age_breakdown,
        gender_breakdown=gender_breakdown,
        start_date=start_date,
        end_date=end_date,
    )


def parse_template_file(filepath: str, platform: str = "MPN", format_type: str = "Banner") -> TemplateMetadata:
    """
    Parse the master template file.
    Currently extracts platform/format from arguments (passed by the user via UI).
    Extend to read booked impressions from the template sheet if needed.
    """
    wb = load_workbook(filepath, data_only=True)

    # Attempt to read from 'MPN & CPN Breakdown' sheet if present
    booked: Dict[str, float] = {}
    if "MPN & CPN Breakdown" in wb.sheetnames:
        # Future: scan the sheet for booked impression values per audience
        pass

    return TemplateMetadata(
        platform=platform,
        format_type=format_type,
        booked_impressions=booked,
    )


# ---------------------------------------------------------------------------
# Internal sheet parsers
# ---------------------------------------------------------------------------

def _v(cell_value, default=0) -> float:
    """Safely coerce a cell value to float."""
    if cell_value is None:
        return float(default)
    try:
        return float(cell_value)
    except (TypeError, ValueError):
        return float(default)


def _is_total_row(value) -> bool:
    return value is not None and "grand total" in str(value).strip().lower()


def _parse_reach_sheet(sheet) -> ReachData:
    """REACH sheet — row 2 contains: impressions, clicks, ctr, reach, frequency."""
    r = 2
    return ReachData(
        actual_impressions=_v(sheet.cell(r, 1).value),
        link_clicks=_v(sheet.cell(r, 2).value),
        ctr=_v(sheet.cell(r, 3).value),
        reach=_v(sheet.cell(r, 4).value),
        frequency=_v(sheet.cell(r, 5).value) or 3.0,
    )


def _parse_device_sheet(sheet) -> List[DeviceBreakdown]:
    devices: List[DeviceBreakdown] = []
    r = 2
    while True:
        name = sheet.cell(r, 1).value
        if name is None:
            break
        if _is_total_row(name):
            r += 1
            continue
        mapped = DEVICE_MAPPING.get(str(name).strip(), str(name).strip())
        devices.append(DeviceBreakdown(
            device_type=mapped,
            impressions=_v(sheet.cell(r, 2).value),
            clicks=_v(sheet.cell(r, 3).value),
            ctr=_v(sheet.cell(r, 4).value),
        ))
        r += 1
    return devices


def _parse_creative_sheet(sheet) -> List[CreativeBreakdown]:
    creatives: List[CreativeBreakdown] = []
    r = 2
    while True:
        name = sheet.cell(r, 1).value
        if name is None:
            break
        if _is_total_row(name):
            r += 1
            continue
        creatives.append(CreativeBreakdown(
            name=str(name).strip(),
            impressions=_v(sheet.cell(r, 2).value),
            clicks=_v(sheet.cell(r, 3).value),
            ctr=_v(sheet.cell(r, 4).value),
        ))
        r += 1
    return creatives


def _parse_age_sheet(sheet) -> List[AgeBreakdown]:
    AGE_BANDS = ["18-24", "25-34", "35-44", "45-54", "55-64", "65+"]
    rows: Dict[str, AgeBreakdown] = {}
    r = 2
    while True:
        name = sheet.cell(r, 1).value
        if name is None:
            break
        if _is_total_row(name):
            r += 1
            continue
        band = str(name).strip()
        rows[band] = AgeBreakdown(
            age_band=band,
            impressions=_v(sheet.cell(r, 2).value),
            clicks=_v(sheet.cell(r, 3).value),
            ctr=_v(sheet.cell(r, 4).value),
        )
        r += 1

    # Return in canonical order; append any extra bands found in file
    result = [rows[b] for b in AGE_BANDS if b in rows]
    for b, v in rows.items():
        if b not in AGE_BANDS:
            result.append(v)
    return result or [AgeBreakdown(b, 0, 0, 0) for b in AGE_BANDS[:4]]


def _parse_gender_sheet(sheet) -> List[GenderBreakdown]:
    GENDERS = ["Male", "Female", "Unknown"]
    rows: Dict[str, GenderBreakdown] = {}
    r = 2
    while True:
        name = sheet.cell(r, 1).value
        if name is None:
            break
        if _is_total_row(name):
            r += 1
            continue
        g = str(name).strip()
        rows[g] = GenderBreakdown(
            gender=g,
            impressions=_v(sheet.cell(r, 2).value),
            clicks=_v(sheet.cell(r, 3).value),
            ctr=_v(sheet.cell(r, 4).value),
        )
        r += 1

    result = [rows[g] for g in GENDERS if g in rows]
    for g, v in rows.items():
        if g not in GENDERS:
            result.append(v)
    return result or [GenderBreakdown("Male", 0, 0, 0), GenderBreakdown("Female", 0, 0, 0)]


def _parse_date_sheet(sheet) -> Tuple[Optional[datetime], Optional[datetime]]:
    dates: List[datetime] = []
    r = 2
    while True:
        val = sheet.cell(r, 1).value
        if val is None:
            break
        if _is_total_row(val):
            r += 1
            continue
        if isinstance(val, datetime):
            dates.append(val)
        elif isinstance(val, str):
            for fmt in ("%d %B, %Y", "%d %B %Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    dates.append(datetime.strptime(val.strip(), fmt))
                    break
                except ValueError:
                    continue
        r += 1

    if not dates:
        today = datetime.today()
        return today, today
    return min(dates), max(dates)
